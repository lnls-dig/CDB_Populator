# -*- coding: utf-8 -*-
"""
Created on Web, Aug 07 ‎15:56:38 2020

CDB Add RFFE uCs

@author: Caio Santos - DIG
"""

import cdb
from cdb.cdb_web_service.api.itemRestApi import ItemRestApi
from cdb.common.exceptions.invalidRequest import InvalidRequest
from cdb.common.exceptions.objectNotFound import ObjectNotFound
import openpyxl
import unidecode
import getpass
import json

print('Please insert the credentials:')
#Reads the credentials and save in the variables to connect to the server
protocol = raw_input("\nNetwork protocol: ")
server = raw_input("CDB server: ")
port = int(input("Server port: "))
user = raw_input("CDB user name: ")
password = getpass.getpass(prompt = "Password: ")

#Log into CDB
login = ItemRestApi(user, password, server, port, protocol)
print('\nLogin successfully.\n\nReading worksheet...\n')

#Reads the Excel workbook file with the data
wb = openpyxl.load_workbook('config_ebpm_total.xlsx')
uc = wb['RFFE_uC']

items = []
x = 0

for i in range(2, 290):
    if uc.cell(row = i, column = 1).value == None:
        continue

    elif uc.cell(row = i, column = 4).value == None:
        continue

    else:
        item = 'RFFE_UC'
        ver = '1.2'
        ipn = uc.cell(row = i, column = 1).value
        sn = uc.cell(row = i, column = 4).value
        obs = uc.cell(row = i, column = 7).value

        name = item+':'+ver+':'+ipn

        if obs == None:
            items.append([name, ipn, sn])
        else:
            obs = unidecode.unidecode(obs)
            obs = 'ATMOS: '+obs
            items.append([name, ipn, sn, obs])

for i in range(0, len(items)):
    try:
        login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '94')
        print('RFFE UC %s exists.') % (items[i][1])

        if len(items[i]) > 3:
            id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '94')[u'id']
            print('%s observations are: %s') % (items[i][1], items[i][3])
            print('Updating observations to %s...\n') % (items[i][1])
            login.addLogEntryToItemWithItemId(id, items[i][3])

        else:
            print("%s doesn't have any observations\n") % (items[i][1])

    except cdb.common.exceptions.objectNotFound.ObjectNotFound:
        print("RFFE UC %s doesn't exist. Uploading it to CDB...") % (items[i][1])
        login.addItem('Inventory', items[i][0], 'Sample', itemIdentifier1 = items[i][2], derivedFromItemId = '94')
        id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '94')[u'id']
        print('Item %s uploaded successfully with ID %s!') % (items[i][0], id)
        x += 1

        if len(items[i]) > 3:
            print('%s observations are: %s') % (items[i][1], items[i][3])
            print('Updating observations to %s...\n') % (items[i][1])
            login.addLogEntryToItemWithItemId(id, items[i][3])

        else:
            print("%s doesn't have any observations\n") % (items[i][1])

print('%s items were uploaded to CDB successfully') % (x)
