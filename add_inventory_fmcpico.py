# -*- coding: utf-8 -*-
"""
Created on Web, Aug 05 ‎13:04:46 2020

CDB Add FMCs Pico IM4

@author: Caio Santos - DIG
"""

import cdb
from cdb.cdb_web_service.api.itemRestApi import ItemRestApi
from cdb.common.exceptions.invalidRequest import InvalidRequest
from cdb.common.exceptions.objectNotFound import ObjectNotFound
import openpyxl
import unidecode
import getpass

print('Please insert the credentials:')
#Reads the credentials and save in the variables to connect to the server
protocol = raw_input("\nNetwork protocol: ")
server = raw_input("CDB server: ")
port = int(input("Server port: "))
user = raw_input("CDB user name: ")
password = getpass.getpass(prompt = "Password: ")

#Log into CDB
login = ItemRestApi(user, password, server, port, protocol)
print('\nLogin successfully.\n\nReading worksheet...\n')

#Reads the Excel workbook with the data
wb = openpyxl.load_workbook('config_ebpm_total.xlsx')
fmc = wb['FMC-Pico-IM4']

items = []
x = 0

for r in range(3, 29):
    item = fmc.cell(r, 1).value
    item = item.split('-')
    item = '_'.join([item[0], item[1], item[2]])

    ipn = fmc.cell(r, 2).value
    ver = fmc.cell(r, 3).value
    sn = fmc.cell(r, 5).value
    name = str(item)+':'+str(ver)+':'+str(ipn)

    items.append([name, ipn, sn])


for i in range(0, 26):
    try:
        login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '66')
        print('FMC Pico IM4 %s exists.') % (items[i][1])

    except cdb.common.exceptions.objectNotFound.ObjectNotFound:
        print("FMC Pico IM4 %s doesn't exist. Uploading it to CDB...") % (items[i][1])
        login.addItem('Inventory', items[i][0], 'Sample', itemIdentifier1 = items[i][2], derivedFromItemId = '66')
        print('Item %s uploaded successfully!') % (items[i][0])
        x += 1

if x > 0:
    print('\n%s items were uploaded to CDB successfully') % (x)

else:
    print('\nYour Database is already updated!')
