# -*- coding: utf-8 -*-
"""
Created on Web, Aug 07 ‎10:42:46 2020

CDB Add FMCs POF 5CH

@author: Caio Santos - DIG
"""

import cdb
from cdb.cdb_web_service.api.itemRestApi import ItemRestApi
from cdb.common.exceptions.invalidRequest import InvalidRequest
from cdb.common.exceptions.objectNotFound import ObjectNotFound
import openpyxl
import unidecode
import getpass

print('Please insert the credentials:')
#Reads the credentials and save in the variables to connect to the server
protocol = raw_input("\nNetwork protocol: ")
server = raw_input("CDB server: ")
port = int(input("Server port: "))
user = raw_input("CDB user name: ")
password = getpass.getpass(prompt = "Password: ")

#Log into CDB
login = ItemRestApi(user, password, server, port, protocol)
print('\nLogin successfully.\n\nReading worksheet...\n')

#Reads the Excel workbook with the data
wb = openpyxl.load_workbook('config_ebpm_total.xlsx')
fmc = wb['FMC 5CH POF']

items = []
x = 0

for r in range(3, 47):
    item = fmc.cell(r, 1).value
    item = item.split(' ')
    item = '_'.join([item[0], item[1], item[2]])

    ipn = fmc.cell(r, 2).value
    ver = fmc.cell(r, 3).value
    sn = fmc.cell(r, 6).value
    obs = fmc.cell(r, 9).value
    name = str(item)+':'+str(ver)+':'+str(ipn)

    if obs != None:
        obs = 'ATMOS: '+unidecode.unidecode(obs)
        items.append([name, ipn, sn, obs])

    else:
        items.append([name, ipn, sn])

for i in range(0, len(items)):
    try:
        login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '79')
        print('FMC 5CH POF %s exists.') % (items[i][1])

        if len(items[i]) > 3:
            id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '79')[u'id']
            print('%s observations are: %s') % (items[i][1], items[i][3])
            print('Updating observations to %s...\n') % (items[i][1])
            login.addLogEntryToItemWithItemId(id, items[i][3])
        else:
            print("%s doesn't have any observations\n") % (items[i][1])

    except cdb.common.exceptions.objectNotFound.ObjectNotFound:
        print("FMC 5CH POF %s doesn't exist. Uploading it to CDB...") % (items[i][1])
        login.addItem('Inventory', items[i][0], 'Sample', itemIdentifier1 = items[i][2], derivedFromItemId = '79')
        print('Item %s uploaded successfully!') % (items[i][0])
        x += 1

        if len(items[i]) > 3:
            id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '79')[u'id']
            print('%s observations are: %s') % (items[i][1], items[i][3])
            print('Updating observations to %s...\n') % (items[i][1])
            login.addLogEntryToItemWithItemId(id, items[i][3])
        else:
            print("%s doesn't have any observations\n") % (items[i][1])

print('%s items were uploaded to CDB successfully') % (x)
