# -*- coding: utf-8 -*-
"""
Created on Tue, May 19 ‎12:14:25 2020

CDB Add Item

@author: Caio Santos - DIG
"""

import cdb
from cdb.cdb_web_service.api.itemRestApi import ItemRestApi
from cdb.common.exceptions.invalidRequest import InvalidRequest
from cdb.common.exceptions.objectNotFound import ObjectNotFound
import openpyxl
import unidecode
import getpass

print('Please insert the credentials:')
#Reads the credentials and save in the variables to connect to the server
protocol = raw_input("\nNetwork protocol: ")
server = raw_input("CDB server: ")
port = int(input("Server port: "))
user = raw_input("CDB user name: ")
password = getpass.getpass(prompt = "Password: ")

#Log into CDB database
login = ItemRestApi(user, password, server, port, protocol)
print('Login successfully.\n\nReading worksheet...\n')

#Opens a Excel sheet to receive the inventory items data and save as a workbook
wb = openpyxl.load_workbook('config_ebpm_total.xlsx')
afc = wb['AFC']

items = []
x = 0

#Reads the Excel and get the useful data
for r in range(4, 200):
    if afc.cell(row = r, column = 1).value == None:
        break

    item = afc.cell(r, 1).value
    ipn = afc.cell(r, 2).value
    ver = afc.cell(r, 3).value
    name = str(item)+':'+str(ver)+':'+str(ipn)
    sn = afc.cell(r, 5).value
    obs = afc.cell(r, 28).value

    if obs == None:
        items.append([name, ipn, sn])

    else:
        obs = unidecode.unidecode(obs)
        obs = 'ATMOS: '+obs
        items.append([name, ipn, sn, obs])

for i in range(0, len(items)):
    if ':3.1:' in  items[i][0]:
        try:
            login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '6')
            print('AFC %s exists.') % (items[i][1])

            if len(items[i]) > 3:
                id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '6')[u'id']
                print('%s observations are: %s') % (items[i][1], items[i][3])
                print('Updating observations to %s...\n') % (items[i][1])
                login.addLogEntryToItemWithItemId(id, items[i][3])

            else:
                print("%s doesn't have any observations\n") % (items[i][1])

        except cdb.common.exceptions.objectNotFound.ObjectNotFound:
            print("AFC %s doesn't exist. Uploading it to CDB...") % (items[i][1])
            login.addItem('Inventory', items[i][0], 'Sample', itemIdentifier1 = items[i][2], derivedFromItemId = '6')
            print('Item %s uploaded successfully!') % (items[i][0])
            x += 1

            if len(items[i]) > 3:
                id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '6')[u'id']
                print('%s observations are: %s') % (items[i][1], items[i][3])
                print('Updating observations to %s...\n') % (items[i][1])
                login.addLogEntryToItemWithItemId(id, items[i][3])

            else:
                print("%s doesn't have any observations\n") % (items[i][1])

    elif ':3.1T:' in items[i][0]:
        try:
            login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '78')
            print('AFC Timing %s exists.') % (items[i][1])

            if len(items[i]) > 3:
                id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '78')[u'id']
                print('%s observations are: %s') % (items[i][1], items[i][3])
                print('Updating observations to %s...\n') % (items[i][1])
                login.addLogEntryToItemWithItemId(id, items[i][3])

            else:
                print("%s doesn't have any observations\n") % (items[i][1])

        except cdb.common.exceptions.objectNotFound.ObjectNotFound:
            print("AFC Timing %s doesn't exist. Uploading it to CDB...") % (items[i][1])
            login.addItem('Inventory', items[i][0], 'Sample', itemIdentifier1 = items[i][2], derivedFromItemId = '78')
            print('Item %s uploaded successfully!') % (items[i][0])
            x += 1

            if len(items[i]) > 3:
                id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '78')[u'id']
                print('%s observations are: %s') % (items[i][1], items[i][3])
                print('Updating observations to %s...\n') % (items[i][1])
                login.addLogEntryToItemWithItemId(id, items[i][3])

            else:
                print("%s doesn't have any observations\n") % (items[i][1])

print('%s items were uploaded to CDB successfully') % (x)
