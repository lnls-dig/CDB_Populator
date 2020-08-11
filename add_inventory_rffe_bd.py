# -*- coding: utf-8 -*-
"""
Created on Web, Aug 10 ‎17:03:02 2020

CDB Add RFFE BDs

@author: Caio Santos - DIG
"""

import cdb
from cdb.cdb_web_service.api.itemRestApi import ItemRestApi
from cdb.common.exceptions.invalidRequest import InvalidRequest
from cdb.common.exceptions.objectNotFound import ObjectNotFound
import openpyxl
import unidecode
import getpass

print('Please insert the credentials:')
#Reads the credentials and save in the variables to connect to the server
protocol = raw_input("\nNetwork protocol: ")
server = raw_input("CDB server: ")
port = int(input("Server port: "))
user = raw_input("CDB user name: ")
password = getpass.getpass(prompt = "Password: ")

#Log into CDB
login = ItemRestApi(user, password, server, port, protocol)
print('\nLogin successfully.\n\nReading worksheet...\n')

#Reads the Excel workbook with the data
wb = openpyxl.load_workbook('config_ebpm_total.xlsx')
bd = wb['RFFE_BD']

items = []
x = 0

for r in range(3, 265):
    if bd.cell(row = r, column = 1).value == None:
        break

    else:
        item = bd.cell(r, 1).value
        ipn = bd.cell(r, 2).value
        ver = '6.0'
        name = item+':'+ver+':'+ipn

        lote = str(bd.cell(r, 6).value)
        ns = int(bd.cell(r, 5).value)

        if ns < 10:
            num = '000'+str(ns)

        elif ns < 100:
            num = '00'+str(ns)

        elif ns >= 100:
            num = '0'+str(ns)

        sn = lote+num
        obs = bd.cell(r, 16).value

        if obs == None:
            items.append([name, ipn, sn])

        else:
            obs = unidecode.unidecode(obs)
            obs = 'ATMOS: '+obs
            items.append([name, ipn, sn, obs])

for i in range(0, len(items)):
    try:
        login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '92')
        print('RFFE BD %s exists.') % (items[i][1])

        if len(items[i]) > 3:
            id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '92')[u'id']
            print('%s observations are: %s') % (items[i][1], items[i][3])
            print('Updating observations to %s...\n') % (items[i][1])
            login.addLogEntryToItemWithItemId(id, items[i][3])

        else:
            print("%s doesn't have any observations\n") % (items[i][1])

    except cdb.common.exceptions.objectNotFound.ObjectNotFound:
        print("RFFE BD %s doesn't exist. Uploading it to CDB...") % (items[i][1])
        login.addItem('Inventory', items[i][0], 'Sample', itemIdentifier1 = items[i][2], derivedFromItemId = '92')
        id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '92')[u'id']
        print('Item %s uploaded successfully with ID %s!') % (items[i][0], id)
        x += 1

        if len(items[i]) > 3:
            print('%s observations are: %s') % (items[i][1], items[i][3])
            print('Updating observations to %s...\n') % (items[i][1])
            login.addLogEntryToItemWithItemId(id, items[i][3])

        else:
            print("%s doesn't have any observations\n") % (items[i][1])

print('%s items were uploaded to CDB successfully') % (x)
