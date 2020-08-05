# -*- coding: utf-8 -*-
"""
Created on Mon, Aug 03 ‎15:54:07 2020

CDB Add Item

@author: Caio Santos - DIG
"""

import cdb
from cdb.cdb_web_service.api.itemRestApi import ItemRestApi
from cdb.common.exceptions.invalidRequest import InvalidRequest
from cdb.common.exceptions.objectNotFound import ObjectNotFound
import openpyxl
import unidecode
import getpass

print('Please insert the credentials:')
#Reads the credentials and save in the variables to connect to the server
protocol = raw_input("\nNetwork protocol: ")
server = raw_input("CDB server: ")
port = int(input("Server port: "))
user = raw_input("CDB user name: ")
password = getpass.getpass(prompt = "Password: ")

#Log into CDB
login = ItemRestApi(user, password, server, port, protocol)
print('Login successfully.\n\nReading worksheet...\n')

#Reads the Excel workbook with the data
wb = openpyxl.load_workbook('config_ebpm_total.xlsx')
fmc = wb['FMC_250M']

items = []
c1 = 1
c2 = 2
c3 = 3
c5 = 5
c26 = 26
x = 0

for r in range(3, 265):
    item = fmc.cell(r, c1).value
    ipn = fmc.cell(r, c2).value
    ver = fmc.cell(r, c3).value
    sn = fmc.cell(r, c5).value
    name = str(item)+':'+str(ver)+':'+str(ipn)
    obs = fmc.cell(r, c26).value

    items.append([name, ipn, sn, obs])

for i in range(0, 257):
    try:
        login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '7')
        print('FMC 250M %s exists.') % (items[i][1])

        if items[i][3] == None:
            print("%s doesn't have any observations") % (items[i][1])

        else:
            items[i][3] = unidecode.unidecode(items[i][3])
            print('%s observations are: %s\n') % (items[i][1], items[i][3])
            print('Updating observations to %s') % (items[i][1])

            id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '7')[u'id']
            obs = 'ATMOS: '+items[i][3]
            print(obs)
            login.addLogEntryToItemWithItemId(id, obs)
            print('Update complete in item %s.\n') % (id)

    except cdb.common.exceptions.objectNotFound.ObjectNotFound:
        print("FMC 250M %s doesn't exist. Uploading it to CDB...") % (items[i][1])
        login.addItem('Inventory', items[i][0], 'Sample', itemIdentifier1 = items[i][2], derivedFromItemId = '7')
        login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '7')
        print('Item %s uploaded successfully!') % (items[i][0])
        x += 1

        if items[i][3] == None:
            print("%s doesn't have any observations\n") % (items[i][1])

        else:
            items[i][3] = unidecode.unidecode(items[i][3])
            print('%s observations are: %s\n') % (items[i][1], items[i][3])
            print('Updating observations to %s') % (items[i][1])

            id = login.getItemByUniqueAttributes('Inventory', items[i][0], itemIdentifier1 = items[i][2], derivedFromItemId = '7')[u'id']
            obs = 'ATMOS: '+items[i][3]
            print(obs)
            login.addLogEntryToItemWithItemId(id, obs)
            print('Update complete in item %s.\n') % (id)

print('%s items were uploaded to CDB successfully') % (x)
